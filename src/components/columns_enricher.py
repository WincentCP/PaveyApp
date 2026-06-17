"""
scrape_places_columns.py
========================
Scraping kolom tambahan dari Google Places API berdasarkan Place ID:
  - place_type       : 'restaurant' atau 'destination'
  - total_reviews    : jumlah total user reviews
  - price_level      : tingkat harga ($ / $$ / $$$ / $$$$) atau 'N/A'

Cara pakai:
  1. Set environment variable GOOGLE_API_KEY, atau isi langsung di bawah.
  2. pip install requests pandas tqdm openpyxl
  3. python scrape_places_columns.py

Output:
  - dataset_rekomendasi_enriched.csv  (CSV utama)
  - dataset_rekomendasi_enriched.xlsx (Excel dengan formatting)
  - scrape_errors.csv                 (baris yang gagal di-fetch, bisa di-retry)
"""

import os
import time
import json
import logging
import requests
import pandas as pd
from tqdm import tqdm

# ─────────────────────────────────────────────
# KONFIGURASI — ganti ini sesuai kebutuhan
# ─────────────────────────────────────────────
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY", "AIzaSyCzghOGFC5qivuKfqjzP5Nm4cvDtlttBG0")

INPUT_CSV      = "dataset_rekomendasi_final.csv"
OUTPUT_CSV     = "dataset_rekomendasi_enriched.csv"
OUTPUT_XLSX    = "dataset_rekomendasi_enriched.xlsx"
ERROR_CSV      = "scrape_errors.csv"

REQUEST_DELAY  = 0.15   # detik antar request (hindari rate limit; Google limit 100 req/detik)
MAX_RETRIES    = 3       # retry otomatis jika timeout/error jaringan

# Keyword untuk klasifikasi restaurant (fallback jika Google types tidak tersedia)
RESTAURANT_KEYWORDS = {
    "restaurant", "food", "cafe", "bar", "bakery", "meal",
    "dining", "buffet", "coffee", "bistro", "eatery", "brunch",
    "lunch", "dinner", "dessert", "pizza", "sushi", "catering",
}

# Google place types yang termasuk restaurant/F&B
GOOGLE_RESTAURANT_TYPES = {
    "restaurant", "food", "cafe", "bar", "bakery",
    "meal_delivery", "meal_takeaway", "night_club",
}

# Mapping price_level integer → label string
PRICE_LEVEL_MAP = {
    0: "Free",
    1: "$",
    2: "$$",
    3: "$$$",
    4: "$$$$",
}

# ─────────────────────────────────────────────
# SETUP LOGGING
# ─────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("scrape.log"),
        logging.StreamHandler(),
    ]
)
logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# FUNGSI HELPER
# ─────────────────────────────────────────────

def classify_from_features(features_text: str) -> str:
    """Klasifikasi fallback berdasarkan kolom features_text yang sudah ada."""
    text = str(features_text).lower()
    for kw in RESTAURANT_KEYWORDS:
        if kw in text:
            return "restaurant"
    return "destination"


def fetch_place_details(place_id: str, session: requests.Session) -> dict:
    """
    Fetch detail dari Google Places API v1 (Place Details).
    Mengembalikan dict dengan keys: place_type, total_reviews, price_level.
    Raise exception jika API error agar bisa di-retry di caller.
    """
    url = "https://maps.googleapis.com/maps/api/place/details/json"
    params = {
        "place_id": place_id,
        "fields": "user_ratings_total,price_level,types",
        "key": GOOGLE_API_KEY,
    }

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = session.get(url, params=params, timeout=10)
            resp.raise_for_status()
            data = resp.json()

            if data.get("status") == "OK":
                result = data.get("result", {})

                # 1. Jumlah review
                total_reviews = result.get("user_ratings_total", None)

                # 2. Tingkat harga
                raw_price = result.get("price_level", None)
                price_level = PRICE_LEVEL_MAP.get(raw_price, "N/A") if raw_price is not None else "N/A"

                # 3. Klasifikasi type dari Google types array
                google_types = set(result.get("types", []))
                if google_types & GOOGLE_RESTAURANT_TYPES:
                    place_type = "restaurant"
                else:
                    place_type = "destination"

                return {
                    "place_type": place_type,
                    "total_reviews": total_reviews,
                    "price_level": price_level,
                    "_api_success": True,
                }

            elif data.get("status") in ("NOT_FOUND", "INVALID_REQUEST"):
                # Bukan error sementara — jangan retry
                logger.warning(f"Place ID tidak ditemukan: {place_id} | Status: {data.get('status')}")
                return {"place_type": None, "total_reviews": None, "price_level": None, "_api_success": False}

            else:
                # Error lain (OVER_QUERY_LIMIT, UNKNOWN_ERROR, dll.) — retry
                logger.warning(f"API status {data.get('status')} untuk {place_id} | Attempt {attempt}/{MAX_RETRIES}")
                time.sleep(2 ** attempt)

        except requests.exceptions.Timeout:
            logger.warning(f"Timeout untuk {place_id} | Attempt {attempt}/{MAX_RETRIES}")
            time.sleep(2 ** attempt)
        except requests.exceptions.RequestException as e:
            logger.warning(f"Request error untuk {place_id}: {e} | Attempt {attempt}/{MAX_RETRIES}")
            time.sleep(2 ** attempt)

    # Semua retry habis
    raise RuntimeError(f"Gagal fetch setelah {MAX_RETRIES} attempt: {place_id}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    if GOOGLE_API_KEY == "ISI_API_KEY_KAMU_DISINI":
        logger.error("❌ GOOGLE_API_KEY belum diset! Set environment variable atau edit script ini.")
        return

    logger.info(f"📂 Membaca dataset: {INPUT_CSV}")
    df = pd.read_csv(INPUT_CSV)
    logger.info(f"   Total baris: {len(df)}")

    # Inisialisasi kolom baru jika belum ada (support resume jika script dihentikan di tengah)
    for col, default in [("place_type", None), ("total_reviews", None), ("price_level", None)]:
        if col not in df.columns:
            df[col] = default

    # Identifikasi baris yang belum diproses (place_type masih None/NaN)
    pending_mask = df["place_type"].isna()
    pending_idx  = df[pending_mask].index.tolist()
    logger.info(f"   Baris yang perlu di-scrape: {len(pending_idx)}")

    if not pending_idx:
        logger.info("✅ Semua baris sudah diproses sebelumnya. Langsung ekspor output.")
    else:
        error_rows = []

        with requests.Session() as session:
            for idx in tqdm(pending_idx, desc="Scraping", unit="row"):
                row = df.loc[idx]
                place_id    = row["id"]
                features    = row.get("features_text", "")

                try:
                    result = fetch_place_details(place_id, session)

                    if result["_api_success"]:
                        df.at[idx, "place_type"]    = result["place_type"]
                        df.at[idx, "total_reviews"] = result["total_reviews"]
                        df.at[idx, "price_level"]   = result["price_level"]
                    else:
                        # API tidak menemukan Place ID — pakai fallback klasifikasi lokal
                        df.at[idx, "place_type"]    = classify_from_features(features)
                        df.at[idx, "total_reviews"] = None
                        df.at[idx, "price_level"]   = "N/A"

                except RuntimeError:
                    # Semua retry habis — tandai sebagai error, isi fallback
                    df.at[idx, "place_type"]    = classify_from_features(features)
                    df.at[idx, "total_reviews"] = None
                    df.at[idx, "price_level"]   = "N/A"
                    error_rows.append({"index": idx, "id": place_id, "name": row["name"]})
                    logger.error(f"ERROR row {idx}: {row['name']} ({place_id})")

                time.sleep(REQUEST_DELAY)

                # Auto-save setiap 100 baris (checkpoint)
                if (pending_idx.index(idx) + 1) % 100 == 0:
                    df.to_csv(OUTPUT_CSV, index=False)
                    logger.info(f"   💾 Checkpoint tersimpan ({pending_idx.index(idx) + 1}/{len(pending_idx)})")

        if error_rows:
            pd.DataFrame(error_rows).to_csv(ERROR_CSV, index=False)
            logger.warning(f"⚠️  {len(error_rows)} baris gagal → disimpan ke {ERROR_CSV}")

    # ─── Fallback lokal untuk baris yang masih None setelah scraping ───
    fallback_mask = df["place_type"].isna()
    if fallback_mask.any():
        df.loc[fallback_mask, "place_type"] = df.loc[fallback_mask, "features_text"].apply(classify_from_features)
        logger.info(f"   Fallback klasifikasi lokal untuk {fallback_mask.sum()} baris.")

    # Konversi total_reviews ke integer (pandas mungkin simpan sebagai float karena NaN)
    df["total_reviews"] = pd.to_numeric(df["total_reviews"], errors="coerce").astype("Int64")

    # ─── Ekspor CSV ───
    df.to_csv(OUTPUT_CSV, index=False)
    logger.info(f"✅ CSV tersimpan: {OUTPUT_CSV}")

    # ─── Ekspor XLSX dengan formatting ───
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Data", index=False)
            ws = writer.sheets["Data"]

            # Header styling
            header_fill = PatternFill("solid", fgColor="1F4E79")
            header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            for cell in ws[1]:
                cell.fill   = header_fill
                cell.font   = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Highlight 3 kolom baru
            new_cols = ["place_type", "total_reviews", "price_level"]
            highlight_fill = PatternFill("solid", fgColor="E2EFDA")
            for col_idx, col_name in enumerate(df.columns, 1):
                if col_name in new_cols:
                    col_letter = get_column_letter(col_idx)
                    for row_idx in range(2, len(df) + 2):
                        ws[f"{col_letter}{row_idx}"].fill = highlight_fill
                        ws[f"{col_letter}{row_idx}"].font = Font(name="Arial", size=10)

            # Auto-fit lebar kolom (max 50 char)
            for col_idx, col_name in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_idx)
                max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max())
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

            # Freeze header row
            ws.freeze_panes = "A2"

        logger.info(f"✅ Excel tersimpan: {OUTPUT_XLSX}")

    except ImportError:
        logger.warning("openpyxl tidak terinstall — hanya ekspor CSV. Jalankan: pip install openpyxl")

    # ─── Ringkasan Statistik ───
    print("\n" + "="*50)
    print("📊 RINGKASAN HASIL SCRAPING")
    print("="*50)
    print(f"Total baris       : {len(df)}")
    print(f"\nDistribusi place_type:")
    print(df["place_type"].value_counts().to_string())
    print(f"\nDistribusi price_level:")
    print(df["price_level"].value_counts().to_string())
    print(f"\nTotal reviews — statistik:")
    print(df["total_reviews"].describe().to_string())
    print(f"\nData kosong (total_reviews): {df['total_reviews'].isna().sum()} baris")
    print("="*50)


if __name__ == "__main__":
    main()
